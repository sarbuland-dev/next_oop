from openpyxl import Workbook
wb=Workbook()
ws=wb.active
name=input("enter your name ")
email=input("enter your email ")
ws["A1"]=name
ws["B1"]=email
ws.append((name,email))
wb.save("creat_file.xlsx")
