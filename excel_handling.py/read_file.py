import os
from openpyxl import load_workbook
wb=load_workbook("creat_file.xlsx")
ws=wb.active
for row in ws.iter_rows(values_only=True):
    print(row)
