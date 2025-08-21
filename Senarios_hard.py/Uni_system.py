from openpyxl import Workbook,load_workbook
import os
import random

class person:
    def __init__(self,name,age,id):
        self.name=name
        self.age=age
        self.id=id
        
        self.file=f"Professor_sheet_{self.name}.xlsx"
        self.teacher_role=[]
        self.st_role=[]
    def show(self):
        return(f"Name: {self.name} Age: {self.age}  ID: {self.id}")

class professor(person):
    def __init__(self,name,age,id,deparment,salary):
        super().__init__(name,age,id)
        self.dep=deparment
        self.salary=salary
    def show2(self):
        # print(f"Name: {self.name} Age: {self.age}  ID: {self.id}  Department: {self.dep}  Salary: {self.salary}")
        base=super().show()
        return( f"{base} Department: {self.dep} Salary: {self.salary}")
    def assign_grade(self):
            
            if os.path.exists(self.file):
                wb=load_workbook(self.file)
                ws=wb.active
            else:
                wb=Workbook()
                ws=wb.active
                ws.append(["Name","Age","ID","department","Salary"])
            ws.append([self.name,self.age,self.id,self.dep,self.salary])
            wb.save(self.file)

class student(person):
    def __init__(self, name, age, id,subject,grade,gpa):
        super().__init__(name, age, id)
        self.sub=subject
        self.grade=grade
        self.gpa=gpa
    
        
        

print("enter your position  ")
print("1: Teacher ")
print("2: Student ")
role=int(input("enter your option "))
if role==1:
    name=input("Enter your name ").capitalize()
age=input("Enter your age  ") 
ran=random.randint(222,999)
first_letter=name[0].upper()
id=f"{first_letter}{ran}"
dep=input("Enter your deparment ") 
salary=input("Enter your salary ")
s2=professor(name,age,id,dep,salary)
print(s2.show2())
print(s2.assign_grade())

          
        
        
# name=input("Enter your name ").capitalize()
# age=input("Enter your age  ") 
# ran=random.randint(222,999)
# first_letter=name[0].upper()
# id=f"{first_letter}{ran}"
# dep=input("Enter your deparment ") 
# salary=input("Enter your salary ")    
# s1=person(name,age,id)
# print(s1.show())
# s2=professor(dep,salary)
# print(s2.show2())
        
        