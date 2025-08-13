from openpyxl import Workbook,load_workbook
import os
import random

class studen_manage_app:
    def __init__(self):
        print(" >>>>> welcome to SB School Attentdence Management System >>>> ")
        # self.file= "student_report.xlsx"
    def main(self):
        print("*"*40)
        teacher_input=input("Enter your name::::  ").capitalize().strip()
        self.file=f"student_report_{teacher_input}.xlsx"
        print("*"*40)
        print(f"Hey! Teacher {teacher_input}, Welcome!")
        print("*"*40)
    
        while True:
         print("*"*40)
         print("Your School Attentdence Menu is here! ")
         print("1: list of students ")
         print("2: Add student  ")
         print("3: Update attentdence of students ")
         print("4: Delete record of student ")
         print("5: Exit ")
         print("*"*40)
         print("*"*40)
         choice=int(input("Enter your choice  "))
         print("*"*40)
         if os.path.exists(self.file):
             wb=load_workbook(self.file)
             ws=wb.active
         else:
             wb=Workbook()
             ws=wb.active
             ws.append(["Name","Rollnumber","Class","Attentdence"])
         wb.save(self.file)
         match choice:
            case 1:
                wb=load_workbook(self.file)
                ws=wb.active
                for row in ws.iter_rows(values_only=True):
                    print("your list of students are ")
                    print("*"*40)
                    print(row)
                    print("*"*40)
            case 2:
                while True:
                  wb=load_workbook(self.file)
                  ws=wb.active
                  print(">>>> Give information about student to add >>>  ")
                  print("*"*40)
                  choice=input("Enter you want to add or not (Add/Not)  ").strip().capitalize()
                  print("*"*40)
                  if choice=="Add":
                     print("*"*40)
                     name=input("Enter student name  ").capitalize().strip()
                     print("*"*40)
                     num=random.randint(2222,9999)
                     first_letter=name[0].upper()
                     rollnum=(f"{first_letter}{num}")
                     print("_"*40)
                     print(f"Your student {name} Rollnumber is {rollnum}")
                     print("_"*40)
                     class_name=input(f"Enter student {name} class  ")
                     print("_"*40)
                     attent=input(f"Enter student {name} attentdence (P/A) ").upper()
                     print("_"*40)
                     print("_"*40)
                     print(f" Your student details are ...... {name},{rollnum},{class_name},{attent}")
                     print("_"*40)
                     print("_"*40)
                     confirm=input("confirm plz   (Yes/No)   ").strip().capitalize()
                     print("_"*40)
                     if confirm=="Yes":
                         ws.append([name,rollnum,class_name,attent])
                         print(">"*40)
                         print("Data save succesfully! ")
                         print(">"*40)
                         wb.save(self.file)
                      
                     elif confirm=="No":
                         print(">"*40)
                         print("Data not save ! ")
                         print(">"*40)
                         self.main()
                         break
                  elif choice=="Not":
                      self.main()  
            case 3: 
                wb=load_workbook(self.file)
                ws=wb.active
                while True:
                 print(">"*40)   
                 choice=input("You want to update attentdence or not (Yes/No)  ").strip().capitalize()
                 print(">"*40)
                 if choice=="Yes":
                      print("_"*40)
                      check_rollnum=input("Enter student rollnumber ") .strip().capitalize()
                      print("_"*40)
                      found_row=None
                      for r in range(2,ws.max_row +1):
                          if (ws.cell(row=r,column=2).value).strip().capitalize()==check_rollnum:
                              found_row=r
                              break
                      if found_row :
                          print("_"*40)
                          new_attent=input("Enter upgraded attentence (P/A)  ").strip().capitalize()
                          print("_"*40)
                          ws.cell(row=found_row,column=4).value=new_attent
                          wb.save(self.file)
                          print("_"*40)
                          print("Updated sucessfully!")
                          print("_"*40)
                      else:
                          print("_"*40)
                          print("Rollnumber not found!!")
                          print("_"*40)
                 elif choice=="No":
                     self.main() 
            case 4:
                wb=load_workbook(self.file)
                ws=wb.active
                while True:
                 print("_"*40)   
                 choice=input("You want to Delete Student record or not (Yes/No)  ").strip().capitalize()
                 print("_"*40)
                 if choice=="Yes":
                      print("_"*40)
                      check_rollnum=input("Enter student rollnumber ") .strip().capitalize()
                      print("_"*40)
                      found_row=None
                      for r in range(2,ws.max_row +1):
                          if (ws.cell(row=r,column=2).value).strip().capitalize()==check_rollnum:
                              found_row=r
                              break
                      if found_row :
                          ws.delete_rows(found_row)
                          wb.save(self.file)
                          print("_"*40)

                          print("Deleted  sucessfully!")
                          print("_"*40)
                      else:
                          print("_"*40)
                          print("Rollnumber not found!!")
                          print("_"*40)
                 elif choice=="No":
                     self.main()
            case 5 :
                print("-"*40)
                print("Thanks for Useing SB School Attentdence Management System ")
                print(f" Goodbye! Teacher.{teacher_input}") 
                print("Exiting......")
                print("-"*40)
                exit()                 
                          
         
                          
                      
                  
                  
                    

                

                
                

        




s1=studen_manage_app()
s1.main()