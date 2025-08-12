import os
from openpyxl import Workbook,load_workbook



class youtube_manage_app:
    def __init__(self):
        print("*"*40)
        print("welcome to Youtube app")
        print("*"*40)
        self.file="youtube.xlsx"
       


    def menu(self):
        print("*"*40)
        print("Enter your details first plz  ")
        print("*"*40)
        name=input("Enter your name  ")
        email=input("Enter your email  ")
        print("*"*40)
        if os.path.exists(self.file):
           wb=load_workbook(self.file)
           ws=wb.active
        else:
           wb=Workbook()
           ws=wb.active
           ws.append(["Name","E-mail","Video name","Time"])
        
        # ws.append((name,email))

        wb.save(self.file)
        # self.current_row=ws.max_row
        print(">"*40)
        print("Login successfully!! ")
        print(">"*40)
        while True:
         print("Youtube menu is here!!")
         print("1: list of videos ")
         print("2: Add a videos ")
         print("3: update a  video ")
         print("4: Delete a videos ")
         print("5: back to menu ")
         print("6: Exit ")
         print("*"*40)
         choice=int(input("Enter your choice "))
         print("*"*40)
         match choice:
            case 1:
                wb=load_workbook(self.file)
                ws=wb.active
                for row in ws.iter_rows(values_only=True):
                    # if row[0] == name and row[1] == email and row[2] != "":
                    
                 print(row)
            case 2:
               print("*"*40)
               video_name=input("Enter video name ")
               video_time=input("Enter video time  ")
               print("*"*40)
               
               wb=load_workbook(self.file)
               ws=wb.active
               ws.append([name,email,video_name,video_time])
               
            
               
               wb.save(self.file)
            case 3:
                    wb = load_workbook(self.file)
                    ws = wb.active
                    print("*"*40)
                    target_video = input("Enter the video name to update: ")
                    print("*"*40)
                    found_row = None
                    for row in range(2, ws.max_row + 1):
                        if (ws.cell(row, 1).value == name and
                            ws.cell(row, 2).value == email and
                            ws.cell(row, 3).value == target_video):
                            found_row = row
                            break

                    if found_row:
                        print(":"*40)
                        new_video = input("Enter new video name: ")
                        new_time = input("Enter new video time: ")
                        print(":"*40)
                        ws.cell(row=found_row, column=3).value = new_video
                        ws.cell(row=found_row, column=4).value = new_time
                        wb.save(self.file)
                        print("*"*40)
                        print("Video updated successfully!")
                        print("*"*40)
                    else:
                        print("*"*40)
                        print("Video not found!")
                        print("*"*40)

            case 4:
                    wb = load_workbook(self.file)
                    ws = wb.active
                    print("*"*40)
                    target_video = input("Enter the video name to delete: ").strip()
                    print("*"*40)
                    found_row = None
                    for row in range(2, ws.max_row + 1):
                        if (
                            ws.cell(row, 3).value == target_video):
                            found_row = row
                            break

                    if found_row:
                        ws.delete_rows(found_row)
                        wb.save(self.file)
                        print("*"*40)
                        print("Video deleted successfully!")
                        print("*"*40)
                    else:
                        print("*"*40)
                        print("Video not found!")
                        print("*"*40)
            case 5:
               self.menu()     
            case 6:
               exit()  


               


               






        
        




s1=youtube_manage_app()
s1.menu()






